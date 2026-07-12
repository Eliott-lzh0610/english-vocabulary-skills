#!/usr/bin/env python3
"""Filter CET-4 word list: remove Gaokao words, keep CET4-exclusive."""

import argparse, os, re, sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from spellchecker import SpellChecker


def extract_first_letter(word: str) -> str:
    w = word.strip()
    if not w:
        return ""
    if w[0] == "(" and len(w) > 1 and w[1].isalpha():
        return w[1].upper()
    if w[0].isalpha():
        return w[0].upper()
    return ""


def filter_exclusive(cet4_dir: str, cet6_dir: str, xlsx_path: str,
                     output_dir: str, freq_threshold: int = 3000):
    spell = SpellChecker()
    _HN = re.compile(r"^(.+?)[123]$")
    _FG = re.compile(r"G(?=[a-z])")

    # Hard-coded K-12 basic words
    _K12 = {
        "january", "february", "march", "april", "may", "june",
        "july", "august", "september", "october", "november", "december",
        "monday", "tuesday", "wednesday", "thursday", "friday",
        "saturday", "sunday", "blackboard", "afternoon", "alphabet",
        "a", "an", "a/an", "the", "i", "you", "he", "she", "it",
        "we", "they", "me", "him", "her", "us", "them",
        "my", "your", "his", "its", "our", "their",
        "am", "is", "are", "was", "were", "be", "been", "being",
        "do", "does", "did", "have", "has", "had",
        "can", "could", "will", "would", "shall", "should",
        "not", "no", "yes", "or", "and", "but", "if", "so",
        "at", "in", "on", "by", "to", "for", "of", "with",
        "from", "up", "down", "out", "off", "over", "under",
        "this", "that", "these", "those", "here", "there",
        "what", "which", "who", "whom", "whose", "when", "where", "why", "how",
        "one", "two", "three", "four", "five", "six", "seven", "eight", "nine", "ten",
        "red", "blue", "green", "yellow", "black", "white",
        "brown", "gray", "grey", "pink", "purple", "orange",
        "mother", "father", "sister", "brother", "uncle", "aunt",
    }

    def _load(d: str) -> set:
        ws = set()
        for fp in Path(d).glob("*_wordlist.txt"):
            with open(fp, encoding="utf-8") as f:
                for line in f:
                    w = line.strip()
                    if w:
                        ws.add(w.lower())
        return ws

    def _load_xlsx(path: str) -> set:
        wb = openpyxl.load_workbook(path, read_only=True)
        all_w = set()
        for sn in wb.sheetnames:
            ws = wb[sn]
            for row in ws.iter_rows(min_row=2, values_only=True):
                for cell in row:
                    if cell and str(cell).strip():
                        raw = str(cell).strip().lower()
                        all_w.add(raw.replace(" ", ""))
                        for p in re.split(r"[\s,/()]+", raw):
                            p = p.strip()
                            if p:
                                all_w.add(p)
        return all_w

    def _normalize(word: str) -> str:
        w = word.lower().strip()
        w = _HN.sub(r"\1", w)
        w = _FG.sub("-", w)
        w = re.sub(r"\(([a-z]+)\)", r"\1", w)
        if "/" in w:
            w = max(w.split("/"), key=len)
        return w

    def _freq(word: str) -> int:
        best = spell.word_frequency[word]
        nf = _normalize(word)
        if nf != word:
            best = max(best, spell.word_frequency[nf])
        if best > 0:
            return best
        if " " not in word and len(word) > 8:
            for sa in ["to", "by", "in", "on", "up", "out", "over", "with",
                       "from", "down", "off", "new", "boom",
                       "board", "room", "book", "man", "men"]:
                idx = word.find(sa)
                if idx > 2:
                    best = max(best, spell.word_frequency[word[:idx] + " " + word[idx:]])
        return best

    c4s = _load(cet4_dir)
    c6s = _load(cet6_dir)
    xlsx_set = _load_xlsx(xlsx_path)

    print(f"  CET4: {len(c4s)} words")
    print(f"  CET6: {len(c6s)} words")
    print(f"  XLSX: {len(xlsx_set)} entries")

    excl = defaultdict(list)
    stats = {"total": 0, "in_xlsx": 0, "k12_hard": 0, "k12_freq": 0, "kept": 0}

    for w in sorted(c4s):
        stats["total"] += 1
        cl = w
        m = _HN.match(cl)
        if m and (m.group(1) in c4s or m.group(1) in c6s):
            cl = m.group(1)
        if "G" in cl:
            fx = _FG.sub("-", cl)
            if fx != cl:
                cl = fx
        nf = _normalize(cl)
        if nf in _K12 or cl.lower() in _K12:
            stats["k12_hard"] += 1
            continue
        if cl in c6s or nf in c6s:
            continue
        # Check xlsx: any part?
        parts = set()
        parts.add(cl)
        parts.add(nf)
        for p in re.split(r"[/()]", cl):
            p = p.strip()
            if p:
                parts.add(p)
        if parts & xlsx_set:
            stats["in_xlsx"] += 1
            continue
        if _freq(cl) > freq_threshold:
            stats["k12_freq"] += 1
            continue
        fc = extract_first_letter(cl)
        excl[fc if fc and "A" <= fc <= "Z" else "#"].append(cl)
        stats["kept"] += 1

    os.makedirs(output_dir, exist_ok=True)
    for lt in sorted(excl):
        ws = sorted(set(excl[lt]))
        with open(os.path.join(output_dir, f"{lt}_wordlist.txt"), "w", encoding="utf-8") as f:
            for w in ws:
                f.write(w + "\n")
        print(f"  {lt}_wordlist.txt ({len(ws)} words)")

    print(f"  Stats: total={stats['total']} in_xlsx={stats['in_xlsx']} "
          f"k12_hard={stats['k12_hard']} k12_freq={stats['k12_freq']} "
          f"kept={stats['kept']}")


def main():
    p = argparse.ArgumentParser(description="Filter CET4 to exclusive words")
    p.add_argument("--cet4-dir", required=True)
    p.add_argument("--cet6-dir", required=True)
    p.add_argument("--xlsx", required=True, help="3500.xlsx path")
    p.add_argument("-o", "--output-dir", default="output/cet4-exclusive")
    p.add_argument("--freq-threshold", type=int, default=3000)
    args = p.parse_args()
    filter_exclusive(args.cet4_dir, args.cet6_dir, args.xlsx,
                     args.output_dir, args.freq_threshold)


if __name__ == "__main__":
    main()
